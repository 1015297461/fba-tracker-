import json
import os
import threading

from .. import pdf_splitter
from ..utils import _now_iso, PROJECT_ROOT


def _build_products_xlsx(products: list, progress_cb=None) -> bytes:
    """
    将产品列表构建成含嵌入主图的 Excel 文件，返回 bytes。
    progress_cb(cur, total)：每张图片完成时回调，可选。
    图片从亚马逊 CDN 并发下载（最多 5 线程），失败时自动重试 2 次，仍失败留空不中断。
    """
    import io
    import time
    import urllib.request
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    IMG_HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer": "https://www.amazon.com/",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    }
    IMG_PX   = 90
    ROW_H    = 70
    COL_A_W  = 14

    def _fetch_image(url: str):
        """下载并缩放图片，最多重试 3 次（含首次），失败返回 None。"""
        if not url:
            return None
        for attempt in range(3):
            try:
                req = urllib.request.Request(url, headers=IMG_HEADERS)
                with urllib.request.urlopen(req, timeout=12) as r:
                    data = r.read()
                pil = PILImage.open(io.BytesIO(data)).convert("RGB")
                pil.thumbnail((IMG_PX, IMG_PX), PILImage.LANCZOS)
                buf = io.BytesIO()
                pil.save(buf, format="JPEG", quality=85)
                buf.seek(0)
                return buf
            except Exception:
                if attempt < 2:
                    time.sleep(1)
        return None

    # 并发下载所有主图，每完成一张触发进度回调
    image_map: dict = {}
    urls = {p["asin"]: p.get("mainImage") or "" for p in products}
    total = len(urls)
    cur = 0
    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(_fetch_image, url): asin for asin, url in urls.items()}
        for fut in as_completed(futures):
            asin = futures[fut]
            try:
                image_map[asin] = fut.result()
            except Exception:
                image_map[asin] = None
            cur += 1
            if progress_cb:
                try:
                    progress_cb(cur, total)
                except Exception:
                    pass

    # 构建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "产品数据"

    HEADERS = ["主图", "ASIN", "标题", "品牌", "价格", "评分", "评论数",
               "可用性", "卖家", "大类排名", "状态", "链接"]
    COL_WIDTHS = [COL_A_W, 14, 50, 16, 10, 8, 10, 14, 20, 14, 8, 50]

    # 表头样式
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # 数据行
    for ri, p in enumerate(products, start=2):
        bsr = p.get("bestSellerRank") or {}
        rank_str = ""
        if bsr.get("mainCategory") and bsr.get("mainRank") is not None:
            rank_str = f"{bsr['mainCategory']} #{bsr['mainRank']}"

        marketplace = p.get("marketplace", "US")
        domain_map = {"US": "amazon.com", "CA": "amazon.ca", "UK": "amazon.co.uk",
                      "DE": "amazon.de", "JP": "amazon.co.jp", "MX": "amazon.com.mx",
                      "FR": "amazon.fr", "IT": "amazon.it", "ES": "amazon.es",
                      "AU": "amazon.com.au", "SG": "amazon.sg", "IN": "amazon.in"}
        domain = domain_map.get(marketplace, "amazon.com")
        link = f"https://www.{domain}/dp/{p['asin']}"

        row_data = [
            "",                                  # A: 图片占位
            p.get("asin", ""),
            p.get("title", ""),
            p.get("brand", ""),
            p.get("price", ""),
            p.get("rating", ""),
            p.get("reviewCount", ""),
            p.get("availability", ""),
            p.get("seller", ""),
            rank_str,
            "成功" if p.get("status") == "success" else "失败",
            link,
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
            if ci == 12:   # 链接列蓝色超链接
                cell.font = Font(color="0563C1", underline="single")
                cell.hyperlink = link

        ws.row_dimensions[ri].height = ROW_H

        # 嵌入主图
        img_buf = image_map.get(p.get("asin", ""))
        if img_buf:
            xl_img = XLImage(img_buf)
            xl_img.width  = IMG_PX
            xl_img.height = IMG_PX
            ws.add_image(xl_img, f"A{ri}")

    # 冻结首行
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportWorker:
    """
    后台常驻线程，轮询 export_jobs 表中 pending 任务并异步处理。
    每 2 秒检查一次队列，支持并发执行（最多 2 个任务同时跑）。
    """

    def __init__(self, state):
        self._state = state
        self._executor = None
        self._active: dict = {}   # jid -> Future
        self._lock = threading.Lock()
        self._thread = threading.Thread(target=self._loop, daemon=True, name="ExportWorker")

    def start(self):
        import concurrent.futures
        self._executor = concurrent.futures.ThreadPoolExecutor(max_workers=2, thread_name_prefix="ExportJob")
        self._thread.start()

    def _loop(self):
        while True:
            try:
                self._tick()
            except Exception as e:
                print(f"[export-worker] tick error: {e}")
            threading.Event().wait(2.0)

    def _tick(self):
        # 清理已完成的 future
        with self._lock:
            done = [jid for jid, fut in self._active.items() if fut.done()]
            for jid in done:
                del self._active[jid]
        # 最多同时跑 2 个
        if len(self._active) >= 2:
            return
        job = self._state.get_pending_export_job()
        if not job:
            return
        jid = job["id"]
        with self._lock:
            if jid in self._active:
                return
            fut = self._executor.submit(self._run_job, job)
            self._active[jid] = fut

    def _run_job(self, job: dict):
        jid = job["id"]
        self._state.update_export_job(jid, status="processing")
        try:
            params = json.loads(job.get("params") or "{}")
            job_type = job["type"]

            if job_type == "scrape_xlsx":
                task_id = params.get("taskId", "")
                asins   = params.get("asins")        # None = 全部
                products = self._state.get_scrape_products(task_id)
                if asins:
                    asin_set = set(asins)
                    products = [p for p in products if p.get("asin") in asin_set]

                def _cb(cur, total):
                    self._state.update_export_job(jid, progress_cur=cur, progress_total=total)

                self._state.update_export_job(jid, progress_total=len(products))
                data = _build_products_xlsx(products, progress_cb=_cb)

                import uuid
                did = uuid.uuid4().hex[:16]
                # 注册到 pdf_splitter 下载注册表（复用现有下载接口）
                tmp_dir = os.path.join(PROJECT_ROOT, "data", "pdf_tmp", "out")
                os.makedirs(tmp_dir, exist_ok=True)
                fname = job.get("file_name") or f"export_{jid[:8]}.xlsx"
                fpath = os.path.join(tmp_dir, f"{did}_{fname}")
                with open(fpath, "wb") as f:
                    f.write(data)
                pdf_splitter._download_registry[did] = fpath

                self._state.update_export_job(
                    jid,
                    status="done",
                    progress_cur=len(products),
                    download_id=did,
                    completed_at=_now_iso(),
                )
            else:
                raise ValueError(f"unknown job type: {job_type}")

        except Exception as e:
            self._state.update_export_job(
                jid,
                status="failed",
                error=str(e),
                completed_at=_now_iso(),
            )
            print(f"[export-worker] job {jid} failed: {e}")
