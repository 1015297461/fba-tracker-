#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FBA Tracker — 局域网协作服务器 v2
============================================
功能：
  1. 提供静态文件服务
  2. /api/products GET/PUT — SQLite 存储，乐观锁，增量写入
  3. /api/login POST — 用户名/密码换 Token
  4. /api/logout POST — 撤销 Token
  5. /api/me GET — 查询当前用户
  6. 首次启动自动从 data/fba-data.json 迁移数据
  7. 首次启动自动创建 data/fba-users.json（默认账号 admin / fba2025）

用法：
  python3 server.py                       # 默认 0.0.0.0:8002
  python3 server.py --port 9000
  python3 server.py --db /path/fba.db     # 自定义数据库路径
  python3 server.py --users /path/u.json  # 自定义用户配置

按 Ctrl+C 停止。
"""

import http.server
import json
import os
import socket
import argparse
import threading
import socketserver
import sqlite3
import secrets
import datetime
import time
import random
import re
from urllib.parse import urlparse, parse_qs, unquote

import rank_fetcher
import product_fetcher
import pdf_splitter


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _now_iso():
    return datetime.datetime.now().isoformat(timespec="seconds")


def get_lan_ip():
    """探测本机局域网 IP，优先返回 192.168/10/172.16-31 段"""
    import subprocess
    candidates = []
    try:
        result = subprocess.run(
            ["ifconfig"], capture_output=True, text=True, timeout=3
        )
        for line in result.stdout.splitlines():
            line = line.strip()
            if line.startswith("inet ") and "127.0.0.1" not in line:
                ip = line.split()[1]
                if ip.startswith("192.168.") or ip.startswith("10.") or \
                   any(ip.startswith(f"172.{i}.") for i in range(16, 32)):
                    candidates.append(ip)
    except Exception:
        pass
    if candidates:
        return candidates[0]
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    except Exception:
        return "127.0.0.1"
    finally:
        s.close()


# ---------------------------------------------------------------------------
# 数据层：SQLite
# ---------------------------------------------------------------------------

class DbState:
    """SQLite 存储，WAL 模式支持并发读，行级写入替代全量 JSON 重写"""

    def __init__(self, db_path):
        self.db_path = db_path
        self.lock = threading.Lock()
        self._init_db()

    def _conn(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=5000")
        return conn

    def _init_db(self):
        with self._conn() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS meta (
                    key   TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                );
                INSERT OR IGNORE INTO meta VALUES ('version', '0');

                CREATE TABLE IF NOT EXISTS products (
                    id            TEXT PRIMARY KEY,
                    name          TEXT,
                    sku           TEXT,
                    category      TEXT,
                    status        TEXT,
                    lead          TEXT,
                    created_at    TEXT,
                    current_stage TEXT,
                    progress      INTEGER DEFAULT 0,
                    fx_rate       REAL    DEFAULT 7.2,
                    stages        TEXT    DEFAULT '{}',
                    logs          TEXT    DEFAULT '[]',
                    variants      TEXT    DEFAULT '[]',
                    sort_order    INTEGER DEFAULT 0,
                    updated_at    TEXT
                );

                CREATE TABLE IF NOT EXISTS audit_log (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id TEXT,
                    user_name  TEXT,
                    action     TEXT,
                    changed_at TEXT DEFAULT (datetime('now'))
                );

                CREATE TABLE IF NOT EXISTS keyword_tasks (
                    id             TEXT PRIMARY KEY,
                    asin           TEXT NOT NULL,
                    marketplace    TEXT NOT NULL,
                    name           TEXT,
                    keywords       TEXT DEFAULT '[]',
                    keyword_notes  TEXT DEFAULT '{}',
                    schedule       TEXT DEFAULT '[0,6,12,18]',
                    enabled        INTEGER DEFAULT 1,
                    created_at     TEXT,
                    last_run_at    TEXT
                );

                CREATE TABLE IF NOT EXISTS rank_snapshots (
                    id           INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id      TEXT,
                    asin         TEXT,
                    marketplace  TEXT,
                    keyword      TEXT,
                    captured_at  TEXT,
                    organic_rank INTEGER,
                    organic_page INTEGER,
                    sponsored    TEXT,
                    status       TEXT,
                    error        TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_snap_task_kw
                    ON rank_snapshots(task_id, keyword, captured_at);

                CREATE TABLE IF NOT EXISTS scrape_tasks (
                    id           TEXT PRIMARY KEY,
                    marketplace  TEXT NOT NULL,
                    total        INTEGER DEFAULT 0,
                    success      INTEGER DEFAULT 0,
                    failed       INTEGER DEFAULT 0,
                    with_reviews INTEGER DEFAULT 0,
                    status       TEXT DEFAULT 'completed',
                    created_at   TEXT
                );

                CREATE TABLE IF NOT EXISTS scrape_products (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id       TEXT NOT NULL,
                    asin          TEXT NOT NULL,
                    marketplace   TEXT,
                    title         TEXT,
                    brand         TEXT,
                    price         TEXT,
                    rating        TEXT,
                    review_count  TEXT,
                    availability  TEXT,
                    bullet_points TEXT,
                    description   TEXT,
                    main_image    TEXT,
                    images        TEXT,
                    aplus_images  TEXT,
                    specifications TEXT,
                    product_details TEXT,
                    categories    TEXT,
                    seller        TEXT,
                    bsr_main_category TEXT,
                    bsr_main_rank     INTEGER,
                    bsr_sub_category  TEXT,
                    bsr_sub_rank      INTEGER,
                    bsr_raw_text      TEXT,
                    customers_say     TEXT,
                    review_images     TEXT,
                    select_to_learn_more TEXT,
                    status        TEXT,
                    error_message TEXT,
                    scraped_at    TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_scrape_prod_task ON scrape_products(task_id);

                CREATE TABLE IF NOT EXISTS review_tasks (
                    id            TEXT PRIMARY KEY,
                    marketplace   TEXT NOT NULL,
                    asins         TEXT DEFAULT '[]',
                    sort_by       TEXT DEFAULT 'recent',
                    filter_star   TEXT,
                    verified_only INTEGER DEFAULT 0,
                    max_pages     INTEGER DEFAULT 3,
                    total         INTEGER DEFAULT 0,
                    new_count     INTEGER DEFAULT 0,
                    status        TEXT DEFAULT 'completed',
                    created_at    TEXT
                );

                CREATE TABLE IF NOT EXISTS review_results (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id       TEXT NOT NULL,
                    asin          TEXT NOT NULL,
                    marketplace   TEXT,
                    review_id     TEXT NOT NULL,
                    rating        REAL,
                    title         TEXT,
                    author        TEXT,
                    date_raw      TEXT,
                    verified      INTEGER DEFAULT 0,
                    body          TEXT,
                    helpful_count INTEGER,
                    images        TEXT DEFAULT '[]',
                    fetched_at    TEXT,
                    UNIQUE(asin, review_id)
                );
                CREATE INDEX IF NOT EXISTS idx_review_res_asin ON review_results(asin);
                CREATE TABLE IF NOT EXISTS export_jobs (
                    id             TEXT PRIMARY KEY,
                    type           TEXT NOT NULL,
                    label          TEXT,
                    params         TEXT DEFAULT '{}',
                    status         TEXT DEFAULT 'pending',
                    progress_cur   INTEGER DEFAULT 0,
                    progress_total INTEGER DEFAULT 0,
                    error          TEXT,
                    download_id    TEXT,
                    file_name      TEXT,
                    created_at     TEXT,
                    completed_at   TEXT
                );
                CREATE TABLE IF NOT EXISTS ai_analysis_tasks (
                    id           TEXT PRIMARY KEY,
                    skill_id     TEXT NOT NULL,
                    asin         TEXT NOT NULL,
                    params       TEXT DEFAULT '{}',
                    status       TEXT DEFAULT 'pending',
                    error        TEXT,
                    summary      TEXT,
                    files        TEXT DEFAULT '[]',
                    created_at   TEXT,
                    completed_at TEXT
                );
            """)
            # 兼容旧版数据库：幂等追加缺失列
            existing_products = {row[1] for row in conn.execute("PRAGMA table_info(products)")}
            if "variants" not in existing_products:
                conn.execute("ALTER TABLE products ADD COLUMN variants TEXT DEFAULT '[]'")
            if "sort_order" not in existing_products:
                conn.execute("ALTER TABLE products ADD COLUMN sort_order INTEGER DEFAULT 0")
                # 按旧排序（创建时间倒序）回填，避免升级后顺序突变
                rows = conn.execute("SELECT id FROM products ORDER BY created_at DESC").fetchall()
                conn.executemany(
                    "UPDATE products SET sort_order=? WHERE id=?",
                    [(i, row["id"]) for i, row in enumerate(rows)],
                )
            existing_tasks = {row[1] for row in conn.execute("PRAGMA table_info(keyword_tasks)")}
            if "keyword_notes" not in existing_tasks:
                conn.execute("ALTER TABLE keyword_tasks ADD COLUMN keyword_notes TEXT DEFAULT '{}'")

    # ---- 内部辅助 ----

    def _get_version(self, conn):
        row = conn.execute("SELECT value FROM meta WHERE key='version'").fetchone()
        return int(row[0]) if row else 0

    def _row_to_product(self, row):
        return {
            "id":           row["id"],
            "name":         row["name"],
            "sku":          row["sku"],
            "category":     row["category"],
            "status":       row["status"],
            "lead":         row["lead"],
            "createdAt":    row["created_at"],
            "currentStage": row["current_stage"],
            "progress":     row["progress"] or 0,
            "fxRate":       row["fx_rate"] or 7.20,
            "stages":       json.loads(row["stages"]   or "{}"),
            "logs":         json.loads(row["logs"]     or "[]"),
            "variants":     json.loads(row["variants"] or "[]"),
        }

    # ---- 公开接口 ----

    def snapshot(self):
        with self._conn() as conn:
            version  = self._get_version(conn)
            rows     = conn.execute(
                "SELECT * FROM products ORDER BY sort_order ASC"
            ).fetchall()
            return {
                "version":  version,
                "products": [self._row_to_product(r) for r in rows],
            }

    def write(self, new_products, base_version, user=None):
        """
        乐观锁写入：base_version 与当前版本不符则返回 (None, conflict_snapshot)。
        成功返回 (new_version, None)。
        """
        with self.lock:
            with self._conn() as conn:
                current = self._get_version(conn)
                if base_version is not None and int(base_version) != current:
                    return None, self.snapshot()

                new_version = current + 1
                conn.execute(
                    "INSERT OR REPLACE INTO meta VALUES ('version', ?)",
                    [str(new_version)],
                )
                now = _now_iso()

                for i, p in enumerate(new_products):
                    conn.execute(
                        """INSERT OR REPLACE INTO products
                           (id, name, sku, category, status, lead, created_at,
                            current_stage, progress, fx_rate, stages, logs, variants, sort_order, updated_at)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        [
                            p.get("id"),
                            p.get("name"),
                            p.get("sku"),
                            p.get("category"),
                            p.get("status"),
                            p.get("lead"),
                            p.get("createdAt"),
                            p.get("currentStage"),
                            p.get("progress", 0),
                            p.get("fxRate", 7.20),
                            json.dumps(p.get("stages",   {}), ensure_ascii=False),
                            json.dumps(p.get("logs",     []), ensure_ascii=False),
                            json.dumps(p.get("variants", []), ensure_ascii=False),
                            i,
                            now,
                        ],
                    )

                # 删除已不在列表中的产品
                if new_products:
                    ids = [p.get("id") for p in new_products if p.get("id")]
                    placeholders = ",".join("?" for _ in ids)
                    conn.execute(
                        f"DELETE FROM products WHERE id NOT IN ({placeholders})", ids
                    )
                else:
                    conn.execute("DELETE FROM products")

                conn.execute(
                    "INSERT INTO audit_log (product_id, user_name, action, changed_at)"
                    " VALUES (?,?,?,?)",
                    ["__batch__", user or "anonymous", f"write v{new_version}", now],
                )
                conn.commit()
                return new_version, None

    # ---- 关键词排名：任务与快照 ----

    def _row_to_task(self, row):
        return {
            "id":           row["id"],
            "asin":         row["asin"],
            "marketplace":  row["marketplace"],
            "name":         row["name"],
            "keywords":     json.loads(row["keywords"]      or "[]"),
            "keywordNotes": json.loads(row["keyword_notes"] or "{}"),
            "schedule":     json.loads(row["schedule"]      or "[]"),
            "enabled":      bool(row["enabled"]),
            "createdAt":    row["created_at"],
            "lastRunAt":    row["last_run_at"],
        }

    def list_rank_tasks(self):
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM keyword_tasks ORDER BY created_at DESC"
            ).fetchall()
            return [self._row_to_task(r) for r in rows]

    def get_rank_task(self, task_id):
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM keyword_tasks WHERE id=?", [task_id]
            ).fetchone()
            return self._row_to_task(row) if row else None

    def upsert_rank_task(self, t):
        tid = t.get("id") or ("kt" + secrets.token_hex(6))
        now = _now_iso()
        with self.lock:
            with self._conn() as conn:
                row = conn.execute(
                    "SELECT created_at, last_run_at FROM keyword_tasks WHERE id=?", [tid]
                ).fetchone()
                created  = row["created_at"]  if row else now
                last_run = row["last_run_at"] if row else None
                conn.execute(
                    """INSERT OR REPLACE INTO keyword_tasks
                       (id, asin, marketplace, name, keywords, keyword_notes,
                        schedule, enabled, created_at, last_run_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    [
                        tid,
                        (t.get("asin") or "").upper().strip(),
                        (t.get("marketplace") or "US").upper(),
                        t.get("name", ""),
                        json.dumps(t.get("keywords", []), ensure_ascii=False),
                        json.dumps(t.get("keywordNotes", {}), ensure_ascii=False),
                        json.dumps(t.get("schedule", [0, 6, 12, 18])),
                        1 if t.get("enabled", True) else 0,
                        created,
                        last_run,
                    ],
                )
                conn.commit()
        return self.get_rank_task(tid)

    def delete_rank_task(self, task_id):
        with self.lock:
            with self._conn() as conn:
                conn.execute("DELETE FROM keyword_tasks WHERE id=?", [task_id])
                conn.execute("DELETE FROM rank_snapshots WHERE task_id=?", [task_id])
                conn.commit()

    def mark_task_run(self, task_id, ts=None):
        with self._conn() as conn:
            conn.execute(
                "UPDATE keyword_tasks SET last_run_at=? WHERE id=?",
                [ts or _now_iso(), task_id],
            )
            conn.commit()

    def add_snapshot(self, task_id, res):
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    """INSERT INTO rank_snapshots
                       (task_id, asin, marketplace, keyword, captured_at,
                        organic_rank, organic_page, sponsored, status, error)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    [
                        task_id, res.get("asin"), res.get("marketplace"),
                        res.get("keyword"), res.get("captured_at"),
                        res.get("organic_rank"), res.get("organic_page"),
                        json.dumps(res.get("sponsored", []), ensure_ascii=False),
                        res.get("status"), res.get("error"),
                    ],
                )
                conn.commit()

    def get_rank_history(self, task_id, keyword=None, limit=3000):
        with self._conn() as conn:
            if keyword:
                rows = conn.execute(
                    """SELECT * FROM rank_snapshots WHERE task_id=? AND keyword=?
                       ORDER BY captured_at ASC LIMIT ?""",
                    [task_id, keyword, limit],
                ).fetchall()
            else:
                rows = conn.execute(
                    """SELECT * FROM rank_snapshots WHERE task_id=?
                       ORDER BY captured_at ASC LIMIT ?""",
                    [task_id, limit],
                ).fetchall()
            return [{
                "id":          r["id"],
                "keyword":     r["keyword"],
                "capturedAt":  r["captured_at"],
                "organicRank": r["organic_rank"],
                "organicPage": r["organic_page"],
                "sponsored":   json.loads(r["sponsored"] or "[]"),
                "status":      r["status"],
                "error":       r["error"],
            } for r in rows]

    # ---- 产品采集：任务与明细 ----

    def _row_to_scrape_task(self, row):
        return {
            "id":          row["id"],
            "marketplace": row["marketplace"],
            "total":       row["total"],
            "success":     row["success"],
            "failed":      row["failed"],
            "withReviews": bool(row["with_reviews"]),
            "status":      row["status"],
            "createdAt":   row["created_at"],
        }

    def create_scrape_task(self, marketplace, total, with_reviews):
        tid = "st" + secrets.token_hex(6)
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    """INSERT INTO scrape_tasks
                       (id, marketplace, total, success, failed, with_reviews, status, created_at)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    [tid, marketplace, total, 0, 0, 1 if with_reviews else 0, "running", _now_iso()],
                )
                conn.commit()
        return tid

    def accumulate_scrape_task(self, task_id, success_delta, failed_delta):
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    "UPDATE scrape_tasks SET success=success+?, failed=failed+?, status='completed' WHERE id=?",
                    [success_delta, failed_delta, task_id],
                )
                conn.commit()

    def list_scrape_tasks(self, limit=100):
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM scrape_tasks ORDER BY created_at DESC LIMIT ?", [limit]
            ).fetchall()
            return [self._row_to_scrape_task(r) for r in rows]

    def save_scrape_products(self, task_id, products):
        now = _now_iso()
        with self.lock:
            with self._conn() as conn:
                for p in products:
                    conn.execute(
                        """INSERT INTO scrape_products
                           (task_id, asin, marketplace, title, brand, price, rating,
                            review_count, availability, bullet_points, description,
                            main_image, images, aplus_images, specifications, product_details,
                            categories, seller, bsr_main_category, bsr_main_rank,
                            bsr_sub_category, bsr_sub_rank, bsr_raw_text, customers_say,
                            review_images, select_to_learn_more, status, error_message, scraped_at)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        [
                            task_id, p.get("asin"), p.get("marketplace"),
                            p.get("title"), p.get("brand"), p.get("price"), p.get("rating"),
                            p.get("review_count"), p.get("availability"),
                            json.dumps(p.get("bullet_points", []), ensure_ascii=False),
                            p.get("description"),
                            p.get("main_image"),
                            json.dumps(p.get("images", []), ensure_ascii=False),
                            json.dumps(p.get("aplus_images", []), ensure_ascii=False),
                            json.dumps(p.get("specifications", {}), ensure_ascii=False),
                            json.dumps(p.get("product_details", {}), ensure_ascii=False),
                            p.get("categories"), p.get("seller"),
                            p.get("bsr_main_category"), p.get("bsr_main_rank"),
                            p.get("bsr_sub_category"), p.get("bsr_sub_rank"), p.get("bsr_raw_text"),
                            p.get("customers_say"),
                            json.dumps(p.get("review_images", []), ensure_ascii=False),
                            json.dumps(p.get("select_to_learn_more", []), ensure_ascii=False),
                            p.get("status"), p.get("error_message"), now,
                        ],
                    )
                conn.commit()

    def _row_to_scrape_product(self, row):
        return {
            "id":            row["id"],
            "asin":          row["asin"],
            "marketplace":   row["marketplace"],
            "title":         row["title"],
            "brand":         row["brand"],
            "price":         row["price"],
            "rating":        row["rating"],
            "reviewCount":   row["review_count"],
            "availability":  row["availability"],
            "bulletPoints":  json.loads(row["bullet_points"] or "[]"),
            "description":   row["description"],
            "mainImage":     row["main_image"],
            "images":        json.loads(row["images"] or "[]"),
            "aplusImages":   json.loads(row["aplus_images"] or "[]"),
            "specifications": json.loads(row["specifications"] or "{}"),
            "productDetails": json.loads(row["product_details"] or "{}"),
            "categories":    row["categories"],
            "seller":        row["seller"],
            "bestSellerRank": {
                "mainCategory": row["bsr_main_category"],
                "mainRank":     row["bsr_main_rank"],
                "subCategory":  row["bsr_sub_category"],
                "subRank":      row["bsr_sub_rank"],
                "rawText":      row["bsr_raw_text"],
            },
            "customerReviews": {
                "customersSay":     row["customers_say"],
                "reviewImages":     json.loads(row["review_images"] or "[]"),
                "selectToLearnMore": json.loads(row["select_to_learn_more"] or "[]"),
            },
            "status":       row["status"],
            "errorMessage": row["error_message"],
            "scrapedAt":    row["scraped_at"],
        }

    def get_scrape_products(self, task_id):
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM scrape_products WHERE task_id=? ORDER BY id ASC", [task_id]
            ).fetchall()
            return [self._row_to_scrape_product(r) for r in rows]

    def delete_scrape_task(self, task_id):
        with self.lock:
            with self._conn() as conn:
                conn.execute("DELETE FROM scrape_tasks WHERE id=?", [task_id])
                conn.execute("DELETE FROM scrape_products WHERE task_id=?", [task_id])
                conn.commit()

    # ---- 评论采集：任务与评论池 ----

    def _row_to_review_task(self, row):
        return {
            "id":           row["id"],
            "marketplace":  row["marketplace"],
            "asins":        json.loads(row["asins"] or "[]"),
            "sortBy":       row["sort_by"],
            "filterStar":   row["filter_star"],
            "verifiedOnly": bool(row["verified_only"]),
            "maxPages":     row["max_pages"],
            "total":        row["total"],
            "newCount":     row["new_count"],
            "status":       row["status"],
            "createdAt":    row["created_at"],
        }

    def create_review_task(self, marketplace, asins, sort_by, filter_star, verified_only, max_pages):
        tid = "rt" + secrets.token_hex(6)
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    """INSERT INTO review_tasks
                       (id, marketplace, asins, sort_by, filter_star, verified_only,
                        max_pages, total, new_count, status, created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    [
                        tid, marketplace, json.dumps(asins, ensure_ascii=False),
                        sort_by, filter_star, 1 if verified_only else 0,
                        max_pages, len(asins), 0, "running", _now_iso(),
                    ],
                )
                conn.commit()
        return tid

    def accumulate_review_task(self, task_id, batch_asins, new_count_delta):
        """合并本批次的 ASIN 到任务的 asins 列表，累加 new_count，并标记完成。"""
        with self.lock:
            with self._conn() as conn:
                row = conn.execute("SELECT asins FROM review_tasks WHERE id=?", [task_id]).fetchone()
                if not row:
                    return
                existing = json.loads(row["asins"] or "[]")
                merged = list(dict.fromkeys(existing + list(batch_asins)))
                conn.execute(
                    "UPDATE review_tasks SET asins=?, total=?, new_count=new_count+?, status='completed' WHERE id=?",
                    [json.dumps(merged, ensure_ascii=False), len(merged), new_count_delta, task_id],
                )
                conn.commit()

    def list_review_tasks(self, limit=100):
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM review_tasks ORDER BY created_at DESC LIMIT ?", [limit]
            ).fetchall()
            return [self._row_to_review_task(r) for r in rows]

    def get_review_task(self, task_id):
        with self._conn() as conn:
            row = conn.execute("SELECT * FROM review_tasks WHERE id=?", [task_id]).fetchone()
            return self._row_to_review_task(row) if row else None

    def save_review_results(self, task_id, asin, marketplace, reviews):
        """INSERT OR IGNORE 写入评论池（按 asin+review_id 去重），返回新增条数。"""
        now = _now_iso()
        inserted = 0
        with self.lock:
            with self._conn() as conn:
                for r in reviews:
                    review_id = r.get("review_id")
                    if not review_id:
                        continue
                    cur = conn.execute(
                        """INSERT OR IGNORE INTO review_results
                           (task_id, asin, marketplace, review_id, rating, title, author,
                            date_raw, verified, body, helpful_count, images, fetched_at)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        [
                            task_id, asin, marketplace, review_id,
                            r.get("rating"), r.get("title"), r.get("author"), r.get("date_raw"),
                            1 if r.get("verified") else 0, r.get("body"), r.get("helpful_count"),
                            json.dumps(r.get("images", []), ensure_ascii=False), now,
                        ],
                    )
                    inserted += cur.rowcount
                conn.commit()
        return inserted

    def _row_to_review_result(self, row):
        return {
            "id":           row["id"],
            "taskId":       row["task_id"],
            "asin":         row["asin"],
            "marketplace":  row["marketplace"],
            "reviewId":     row["review_id"],
            "rating":       row["rating"],
            "title":        row["title"],
            "author":       row["author"],
            "dateRaw":      row["date_raw"],
            "verified":     bool(row["verified"]),
            "body":         row["body"],
            "helpfulCount": row["helpful_count"],
            "images":       json.loads(row["images"] or "[]"),
            "fetchedAt":    row["fetched_at"],
        }

    def get_review_results(self, task_id):
        with self._conn() as conn:
            task_row = conn.execute("SELECT asins FROM review_tasks WHERE id=?", [task_id]).fetchone()
            if not task_row:
                return []
            asins = json.loads(task_row["asins"] or "[]")
            if not asins:
                return []
            placeholders = ",".join("?" for _ in asins)
            rows = conn.execute(
                f"SELECT * FROM review_results WHERE asin IN ({placeholders}) ORDER BY asin, fetched_at DESC",
                asins,
            ).fetchall()
            return [self._row_to_review_result(r) for r in rows]

    def delete_review_task(self, task_id):
        with self.lock:
            with self._conn() as conn:
                conn.execute("DELETE FROM review_tasks WHERE id=?", [task_id])
                conn.commit()

    # ---- 导出任务 ----

    def create_export_job(self, job_type: str, label: str, params: dict, progress_total: int) -> str:
        import secrets
        jid = "ej" + secrets.token_hex(8)
        now = _now_iso()
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    """INSERT INTO export_jobs
                       (id, type, label, params, status, progress_cur, progress_total, created_at)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    [jid, job_type, label, json.dumps(params, ensure_ascii=False),
                     "pending", 0, progress_total, now]
                )
                conn.commit()
        return jid

    def update_export_job(self, jid: str, **kwargs):
        allowed = {"status", "progress_cur", "progress_total", "error",
                   "download_id", "file_name", "completed_at"}
        fields = {k: v for k, v in kwargs.items() if k in allowed}
        if not fields:
            return
        sets = ", ".join(f"{k}=?" for k in fields)
        vals = list(fields.values()) + [jid]
        with self.lock:
            with self._conn() as conn:
                conn.execute(f"UPDATE export_jobs SET {sets} WHERE id=?", vals)
                conn.commit()

    def list_export_jobs(self, limit: int = 100) -> list:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM export_jobs ORDER BY created_at DESC LIMIT ?", [limit]
            ).fetchall()
        return [dict(r) for r in rows]

    def get_pending_export_job(self):
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM export_jobs WHERE status='pending' ORDER BY created_at ASC LIMIT 1"
            ).fetchone()
        return dict(row) if row else None

    def delete_export_job(self, jid: str):
        with self.lock:
            with self._conn() as conn:
                conn.execute("DELETE FROM export_jobs WHERE id=?", [jid])
                conn.commit()

    # ---- AI分析任务 ----

    def create_ai_task(self, skill_id: str, asin: str, params: dict) -> str:
        tid = "ai" + secrets.token_hex(6)
        with self.lock:
            with self._conn() as conn:
                conn.execute(
                    """INSERT INTO ai_analysis_tasks
                       (id, skill_id, asin, params, status, created_at)
                       VALUES (?,?,?,?,?,?)""",
                    [tid, skill_id, asin, json.dumps(params, ensure_ascii=False),
                     "pending", _now_iso()],
                )
                conn.commit()
        return tid

    def update_ai_task(self, tid: str, **kwargs):
        allowed = {"status", "error", "summary", "files", "completed_at"}
        fields = {k: v for k, v in kwargs.items() if k in allowed}
        if not fields:
            return
        if "files" in fields and not isinstance(fields["files"], str):
            fields["files"] = json.dumps(fields["files"], ensure_ascii=False)
        sets = ", ".join(f"{k}=?" for k in fields)
        vals = list(fields.values()) + [tid]
        with self.lock:
            with self._conn() as conn:
                conn.execute(f"UPDATE ai_analysis_tasks SET {sets} WHERE id=?", vals)
                conn.commit()

    def _row_to_ai_task(self, row):
        d = dict(row)
        d["params"] = json.loads(d.get("params") or "{}")
        d["files"] = json.loads(d.get("files") or "[]")
        return d

    def list_ai_tasks(self, skill_id: str = None, limit: int = 100) -> list:
        with self._conn() as conn:
            if skill_id:
                rows = conn.execute(
                    "SELECT * FROM ai_analysis_tasks WHERE skill_id=? ORDER BY created_at DESC LIMIT ?",
                    [skill_id, limit],
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM ai_analysis_tasks ORDER BY created_at DESC LIMIT ?", [limit]
                ).fetchall()
        return [self._row_to_ai_task(r) for r in rows]

    def get_ai_task(self, tid: str):
        with self._conn() as conn:
            row = conn.execute("SELECT * FROM ai_analysis_tasks WHERE id=?", [tid]).fetchone()
        return self._row_to_ai_task(row) if row else None

    def get_pending_ai_task(self):
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM ai_analysis_tasks WHERE status='pending' ORDER BY created_at ASC LIMIT 1"
            ).fetchone()
        return self._row_to_ai_task(row) if row else None

    def delete_ai_task(self, tid: str):
        with self.lock:
            with self._conn() as conn:
                conn.execute("DELETE FROM ai_analysis_tasks WHERE id=?", [tid])
                conn.commit()

    def import_from_json(self, json_path):
        """首次启动时从旧版 fba-data.json 一键迁移"""
        if not os.path.exists(json_path):
            return False
        with self._conn() as conn:
            count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
            if count > 0:
                return False  # DB 已有数据，跳过

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if isinstance(obj, dict):
                products = obj.get("products", [])
                orig_version = int(obj.get("version", 0))
            elif isinstance(obj, list):
                products = obj
                orig_version = 0
            else:
                return False

            if not products:
                return False

            # 写入后将版本号还原为原始值（避免版本跳变）
            self.write(products, None, user="__migration__")
            with self._conn() as conn:
                conn.execute(
                    "INSERT OR REPLACE INTO meta VALUES ('version', ?)",
                    [str(orig_version)],
                )
                conn.commit()
            print(f"[info] 已从 {json_path} 迁移 {len(products)} 条产品 (v{orig_version})")
            return True
        except Exception as e:
            print(f"[warn] JSON 迁移失败: {e}")
            return False

    @property
    def version(self):
        with self._conn() as conn:
            return self._get_version(conn)

    @property
    def product_count(self):
        with self._conn() as conn:
            return conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]


# ---------------------------------------------------------------------------
# 认证层：简单 Token（局域网场景）
# ---------------------------------------------------------------------------

class AuthManager:
    """
    基于随机 Token 的轻量认证。
    - 用户名/密码存储在 fba-users.json，明文（局域网内可接受）
    - Token 仅在内存中，服务器重启后需重新登录
    - 首次启动自动创建默认账号 admin / fba2025
    """

    def __init__(self, users_path):
        self.users_path = users_path
        self.tokens = {}        # token -> {username, name, role}
        self.lock   = threading.Lock()
        self._ensure_default_users()

    def _ensure_default_users(self):
        if os.path.exists(self.users_path):
            return
        default = {
            "users": [
                {"username": "admin",  "password": "fba2025", "name": "管理员", "role": "admin"},
                {"username": "editor", "password": "fba2025", "name": "编辑员", "role": "editor"},
            ]
        }
        with open(self.users_path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        print(f"[info] 已创建用户配置: {self.users_path}")
        print("  默认账号: admin / fba2025  ← 请尽快修改密码")

    def _load_users(self):
        try:
            with open(self.users_path, "r", encoding="utf-8") as f:
                return json.load(f).get("users", [])
        except Exception:
            return []

    def login(self, username, password):
        """验证成功返回 (token, user_info)，失败返回 (None, None)"""
        for u in self._load_users():
            if u["username"] == username and u["password"] == password:
                token = secrets.token_hex(32)
                info  = {
                    "username": username,
                    "name":     u.get("name", username),
                    "role":     u.get("role", "editor"),
                }
                with self.lock:
                    self.tokens[token] = info
                return token, info
        return None, None

    def verify(self, token):
        """返回 user_info 或 None"""
        if not token:
            return None
        with self.lock:
            return self.tokens.get(token)

    def logout(self, token):
        with self.lock:
            self.tokens.pop(token, None)


# ---------------------------------------------------------------------------
# HTTP Handler
# ---------------------------------------------------------------------------

def run_rank_task(state, task, delay_between_kw=(3.0, 7.0)):
    """串行采集任务下所有关键词，写入快照，返回结果列表。"""
    results = []
    kws = task.get("keywords", [])
    for i, kw in enumerate(kws):
        kw = (kw or "").strip()
        if not kw:
            continue
        try:
            res = rank_fetcher.locate_rank(task["asin"], task["marketplace"], kw, max_pages=3)
        except Exception as e:
            res = {
                "asin": task["asin"], "keyword": kw,
                "marketplace": task["marketplace"], "status": "error",
                "organic_rank": None, "organic_page": None, "sponsored": [],
                "error": f"fetch_exc:{type(e).__name__}", "captured_at": _now_iso(),
            }
        state.add_snapshot(task["id"], res)
        results.append(res)
        if i < len(kws) - 1:
            time.sleep(random.uniform(*delay_between_kw))
    state.mark_task_run(task["id"])
    return results


def run_scrape_task(state, asins, marketplace, with_reviews, task_id=None, total=None):
    """批量采集产品详情，落库并返回 (task_id, results)。
    传入已有 task_id 时，结果累加到该任务，使同一次提交的多个批次合并为一条历史记录。"""
    asins = [a.strip().upper() for a in asins if a and a.strip()]
    if task_id is None:
        task_id = state.create_scrape_task(marketplace, total if total is not None else len(asins), with_reviews)
    try:
        results = product_fetcher.scrape_products(asins, marketplace, with_reviews=with_reviews)
    except Exception as e:
        results = [
            {**product_fetcher._empty_product(a, marketplace, f"fetch_exc:{type(e).__name__}")}
            for a in asins
        ]
    state.save_scrape_products(task_id, results)
    success = sum(1 for r in results if r.get("status") == "success")
    failed = len(results) - success
    state.accumulate_scrape_task(task_id, success, failed)
    return task_id, results


def run_review_task(state, asins, marketplace, sort_by, filter_star, verified_only, max_pages, task_id=None):
    """批量采集评论，去重落入评论池并返回 (task_id, results)。
    传入已有 task_id 时，结果累加到该任务，使同一次提交的多个批次合并为一条历史记录。"""
    asins = [a.strip().upper() for a in asins if a and a.strip()]
    if task_id is None:
        task_id = state.create_review_task(marketplace, asins, sort_by, filter_star, verified_only, max_pages)

    try:
        fetch_results = product_fetcher.fetch_reviews_for_asins(
            asins, marketplace, max_pages=max_pages, sort_by=sort_by,
            filter_by_star=filter_star, verified_only=verified_only,
        )
    except Exception as e:
        fetch_results = [
            {"asin": a, "marketplace": marketplace, "reviews": [],
             "status": "failed", "error_message": f"fetch_exc:{type(e).__name__}"}
            for a in asins
        ]

    new_count = 0
    summaries = []
    for r in fetch_results:
        inserted = state.save_review_results(task_id, r["asin"], marketplace, r["reviews"])
        new_count += inserted
        summaries.append({
            "asin": r["asin"],
            "marketplace": marketplace,
            "totalFetched": len(r["reviews"]),
            "newCount": inserted,
            "status": r["status"],
            "errorMessage": r["error_message"],
        })

    state.accumulate_review_task(task_id, asins, new_count)
    return task_id, summaries


def start_scheduler(state):
    """后台守护线程：每分钟检查定时档位，命中即采集（同一小时不重复）。"""
    def loop():
        done = set()  # (task_id, 'YYYY-MM-DD-HH')，避免同一小时重复触发
        while True:
            try:
                now  = datetime.datetime.now()
                slot = now.strftime("%Y-%m-%d-%H")
                today = now.strftime("%Y-%m-%d")
                for task in state.list_rank_tasks():
                    if not task["enabled"] or not task["keywords"]:
                        continue
                    hours = {h % 24 for h in task["schedule"]}  # 24:00 视作 0:00
                    if now.hour not in hours:
                        continue
                    key = (task["id"], slot)
                    if key in done:
                        continue
                    done.add(key)
                    print(f"[rank] 定时执行 {task['asin']}/{task['marketplace']} "
                          f"({len(task['keywords'])} 词) @ {slot}")
                    try:
                        run_rank_task(state, task)
                    except Exception as e:
                        print(f"[rank] 任务异常: {e}")
                if len(done) > 800:
                    done = {k for k in done if k[1] >= today}
            except Exception as e:
                print(f"[rank] 调度异常: {e}")
            time.sleep(60)

    th = threading.Thread(target=loop, daemon=True)
    th.start()
    print("[rank] 排名调度线程已启动（每分钟检查 0/6/12/18 档位）")


def _extract_token(handler):
    auth = handler.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:].strip()
    return None


def _build_products_xlsx(products: list, progress_cb=None) -> bytes:
    """
    将产品列表构建成含嵌入主图的 Excel 文件，返回 bytes。
    progress_cb(cur, total)：每张图片完成时回调，可选。
    图片从亚马逊 CDN 并发下载（最多 5 线程），失败时自动重试 2 次，仍失败留空不中断。
    """
    import io
    import time
    import urllib.request
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    IMG_HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer": "https://www.amazon.com/",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    }
    IMG_PX   = 90
    ROW_H    = 70
    COL_A_W  = 14

    def _fetch_image(url: str):
        """下载并缩放图片，最多重试 3 次（含首次），失败返回 None。"""
        if not url:
            return None
        for attempt in range(3):
            try:
                req = urllib.request.Request(url, headers=IMG_HEADERS)
                with urllib.request.urlopen(req, timeout=12) as r:
                    data = r.read()
                pil = PILImage.open(io.BytesIO(data)).convert("RGB")
                pil.thumbnail((IMG_PX, IMG_PX), PILImage.LANCZOS)
                buf = io.BytesIO()
                pil.save(buf, format="JPEG", quality=85)
                buf.seek(0)
                return buf
            except Exception:
                if attempt < 2:
                    time.sleep(1)
        return None

    # 并发下载所有主图，每完成一张触发进度回调
    image_map: dict = {}
    urls = {p["asin"]: p.get("mainImage") or "" for p in products}
    total = len(urls)
    cur = 0
    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(_fetch_image, url): asin for asin, url in urls.items()}
        for fut in as_completed(futures):
            asin = futures[fut]
            try:
                image_map[asin] = fut.result()
            except Exception:
                image_map[asin] = None
            cur += 1
            if progress_cb:
                try:
                    progress_cb(cur, total)
                except Exception:
                    pass

    # 构建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "产品数据"

    HEADERS = ["主图", "ASIN", "标题", "品牌", "价格", "评分", "评论数",
               "可用性", "卖家", "大类排名", "状态", "链接"]
    COL_WIDTHS = [COL_A_W, 14, 50, 16, 10, 8, 10, 14, 20, 14, 8, 50]

    # 表头样式
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # 数据行
    for ri, p in enumerate(products, start=2):
        bsr = p.get("bestSellerRank") or {}
        rank_str = ""
        if bsr.get("mainCategory") and bsr.get("mainRank") is not None:
            rank_str = f"{bsr['mainCategory']} #{bsr['mainRank']}"

        marketplace = p.get("marketplace", "US")
        domain_map = {"US": "amazon.com", "CA": "amazon.ca", "UK": "amazon.co.uk",
                      "DE": "amazon.de", "JP": "amazon.co.jp", "MX": "amazon.com.mx",
                      "FR": "amazon.fr", "IT": "amazon.it", "ES": "amazon.es",
                      "AU": "amazon.com.au", "SG": "amazon.sg", "IN": "amazon.in"}
        domain = domain_map.get(marketplace, "amazon.com")
        link = f"https://www.{domain}/dp/{p['asin']}"

        row_data = [
            "",                                  # A: 图片占位
            p.get("asin", ""),
            p.get("title", ""),
            p.get("brand", ""),
            p.get("price", ""),
            p.get("rating", ""),
            p.get("reviewCount", ""),
            p.get("availability", ""),
            p.get("seller", ""),
            rank_str,
            "成功" if p.get("status") == "success" else "失败",
            link,
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
            if ci == 12:   # 链接列蓝色超链接
                cell.font = Font(color="0563C1", underline="single")
                cell.hyperlink = link

        ws.row_dimensions[ri].height = ROW_H

        # 嵌入主图
        img_buf = image_map.get(p.get("asin", ""))
        if img_buf:
            xl_img = XLImage(img_buf)
            xl_img.width  = IMG_PX
            xl_img.height = IMG_PX
            ws.add_image(xl_img, f"A{ri}")

    # 冻结首行
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportWorker:
    """
    后台常驻线程，轮询 export_jobs 表中 pending 任务并异步处理。
    每 2 秒检查一次队列，支持并发执行（最多 2 个任务同时跑）。
    """

    def __init__(self, state):
        self._state = state
        self._executor = None
        self._active: dict = {}   # jid -> Future
        self._lock = threading.Lock()
        self._thread = threading.Thread(target=self._loop, daemon=True, name="ExportWorker")

    def start(self):
        import concurrent.futures
        self._executor = concurrent.futures.ThreadPoolExecutor(max_workers=2, thread_name_prefix="ExportJob")
        self._thread.start()

    def _loop(self):
        while True:
            try:
                self._tick()
            except Exception as e:
                print(f"[export-worker] tick error: {e}")
            threading.Event().wait(2.0)

    def _tick(self):
        # 清理已完成的 future
        with self._lock:
            done = [jid for jid, fut in self._active.items() if fut.done()]
            for jid in done:
                del self._active[jid]
        # 最多同时跑 2 个
        if len(self._active) >= 2:
            return
        job = self._state.get_pending_export_job()
        if not job:
            return
        jid = job["id"]
        with self._lock:
            if jid in self._active:
                return
            fut = self._executor.submit(self._run_job, job)
            self._active[jid] = fut

    def _run_job(self, job: dict):
        jid = job["id"]
        self._state.update_export_job(jid, status="processing")
        try:
            params = json.loads(job.get("params") or "{}")
            job_type = job["type"]

            if job_type == "scrape_xlsx":
                task_id = params.get("taskId", "")
                asins   = params.get("asins")        # None = 全部
                products = self._state.get_scrape_products(task_id)
                if asins:
                    asin_set = set(asins)
                    products = [p for p in products if p.get("asin") in asin_set]

                def _cb(cur, total):
                    self._state.update_export_job(jid, progress_cur=cur, progress_total=total)

                self._state.update_export_job(jid, progress_total=len(products))
                data = _build_products_xlsx(products, progress_cb=_cb)

                import uuid
                did = uuid.uuid4().hex[:16]
                # 注册到 pdf_splitter 下载注册表（复用现有下载接口）
                import tempfile, os
                tmp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "pdf_tmp", "out")
                os.makedirs(tmp_dir, exist_ok=True)
                fname = job.get("file_name") or f"export_{jid[:8]}.xlsx"
                fpath = os.path.join(tmp_dir, f"{did}_{fname}")
                with open(fpath, "wb") as f:
                    f.write(data)
                pdf_splitter._download_registry[did] = fpath

                self._state.update_export_job(
                    jid,
                    status="done",
                    progress_cur=len(products),
                    download_id=did,
                    completed_at=_now_iso(),
                )
            else:
                raise ValueError(f"unknown job type: {job_type}")

        except Exception as e:
            self._state.update_export_job(
                jid,
                status="failed",
                error=str(e),
                completed_at=_now_iso(),
            )
            print(f"[export-worker] job {jid} failed: {e}")


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 无人值守跑 `claude -p` 时收窄到 SKILL.md 自己声明允许使用的工具，
# 不用 --dangerously-skip-permissions / bypassPermissions 整体放开。
AI_ALLOWED_TOOLS = "Bash Read Write Edit mcp__playwright__* mcp__sif-mcp__*"

# 单次分析的预算上限（美元）。cosmo-diagnose 一次完整分析涉及多次 Playwright
# 抓取 + SIF 数据调用 + Alexa 反查 + 报告撰写，真实成本待首次端到端联调后
# 用返回的 total_cost_usd 校准，这里先给一个偏宽松的初始值。
AI_MAX_BUDGET_USD = "8"

AI_SKILLS = {
    "cosmo-diagnose": {
        "label": "Listing诊断（COSMO）",
        "dir": os.path.join(BASE_DIR, "skills", "CosmoDiagnose"),
        "build_prompt": lambda asin, params, task_id: (
            f"使用 cosmo-diagnose skill 分析 ASIN {asin}，"
            f"Alexa 只问前 {int(params.get('alexaQuestions', 3) or 3)} 个问题。"
            f"完成后请把 Phase 3 的四个输出文件保存到当前技能目录下的 "
            f"report/{task_id}/ 子目录（而不是 report/ 根目录），文件命名规则不变。"
        ),
    },
}


class AiAnalysisWorker:
    """
    后台常驻线程，轮询 ai_analysis_tasks 表中 pending 任务，
    串行（最多1个并发）shell 出去跑 `claude -p` headless 会话执行 Skill 分析。
    不复用 ExportWorker：职责不同（本地计算 vs 调起一次可能耗时数分钟、
    真花钱的 Claude Code 会话），且并发上限也不同（这里强制串行）。
    """

    def __init__(self, state):
        self._state = state
        self._executor = None
        self._active: dict = {}   # task_id -> Future
        self._lock = threading.Lock()
        self._thread = threading.Thread(target=self._loop, daemon=True, name="AiAnalysisWorker")

    def start(self):
        import concurrent.futures
        self._executor = concurrent.futures.ThreadPoolExecutor(max_workers=1, thread_name_prefix="AiAnalysisJob")
        self._thread.start()

    def _loop(self):
        while True:
            try:
                self._tick()
            except Exception as e:
                print(f"[ai-worker] tick error: {e}")
            threading.Event().wait(3.0)

    def _tick(self):
        with self._lock:
            done = [tid for tid, fut in self._active.items() if fut.done()]
            for tid in done:
                del self._active[tid]
        if len(self._active) >= 1:
            return
        task = self._state.get_pending_ai_task()
        if not task:
            return
        tid = task["id"]
        with self._lock:
            if tid in self._active:
                return
            self._state.update_ai_task(tid, status="running")
            fut = self._executor.submit(self._run_job, task)
            self._active[tid] = fut

    def _run_job(self, task: dict):
        import subprocess
        tid = task["id"]
        skill = AI_SKILLS.get(task["skill_id"])
        if not skill:
            self._state.update_ai_task(
                tid, status="failed", error=f"unknown skillId: {task['skill_id']}",
                completed_at=_now_iso(),
            )
            return

        asin = task["asin"]
        params = task.get("params") or {}
        prompt = skill["build_prompt"](asin, params, tid)
        skill_dir = skill["dir"]
        out_dir = os.path.join(skill_dir, "report", tid)

        try:
            proc = subprocess.run(
                ["claude", "-p", prompt,
                 "--output-format", "json",
                 "--allowedTools", AI_ALLOWED_TOOLS,
                 "--max-budget-usd", AI_MAX_BUDGET_USD],
                cwd=skill_dir,
                capture_output=True, text=True, timeout=1800,
            )
        except subprocess.TimeoutExpired:
            self._state.update_ai_task(
                tid, status="failed", error="分析超时（超过30分钟）", completed_at=_now_iso(),
            )
            print(f"[ai-worker] task {tid} timeout")
            return
        except Exception as e:
            self._state.update_ai_task(tid, status="failed", error=str(e), completed_at=_now_iso())
            print(f"[ai-worker] task {tid} failed to launch: {e}")
            return

        files = []
        if os.path.isdir(out_dir):
            for fn in sorted(os.listdir(out_dir)):
                ext = fn.rsplit(".", 1)[-1].lower() if "." in fn else ""
                if ext in ("html", "md"):
                    files.append({"name": fn, "type": ext})

        result_obj = None
        try:
            result_obj = json.loads(proc.stdout)
        except (json.JSONDecodeError, TypeError):
            pass

        is_error = bool(result_obj.get("is_error")) if result_obj else (proc.returncode != 0)
        denials = (result_obj or {}).get("permission_denials") or []

        if is_error or not files:
            err = (result_obj or {}).get("result") or (proc.stderr or "")[-2000:] or "分析未产出任何文件"
            self._state.update_ai_task(
                tid, status="failed", error=err[:2000], files=files, completed_at=_now_iso(),
            )
            print(f"[ai-worker] task {tid} failed, returncode={proc.returncode}")
            return

        summary = ((result_obj or {}).get("result") or "")[:500]
        if denials:
            summary += f"\n⚠ 有 {len(denials)} 个工具调用被权限拦截，报告内容可能不完整。"

        cost = (result_obj or {}).get("total_cost_usd") or 0
        self._state.update_ai_task(
            tid, status="done", summary=summary, files=files, completed_at=_now_iso(),
        )
        print(f"[ai-worker] task {tid} done, {len(files)} files, cost=${cost:.4f}")


def make_handler(state, auth):
    class Handler(http.server.SimpleHTTPRequestHandler):

        def end_headers(self):
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
            self.send_header("Pragma", "no-cache")
            super().end_headers()

        def log_message(self, fmt, *args):
            path = args[0] if args else ""
            if "/api/" in str(path):
                return
            print(f"  {self.address_string()} — {path}")

        # ---- GET ----

        def do_GET(self):
            path = urlparse(self.path).path

            if path == "/api/products":
                user = auth.verify(_extract_token(self))
                if not user:
                    self._send_json(401, {"error": "请先登录"})
                    return
                self._send_json(200, state.snapshot())
                return

            if path == "/api/me":
                user = auth.verify(_extract_token(self))
                if not user:
                    self._send_json(401, {"error": "未登录"})
                    return
                self._send_json(200, user)
                return

            if path == "/api/rank/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                self._send_json(200, {"tasks": state.list_rank_tasks()})
                return

            if path == "/api/rank/history":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("taskId") or [""])[0]
                keyword = (q.get("keyword") or [None])[0]
                if not task_id:
                    self._send_json(400, {"error": "taskId required"})
                    return
                self._send_json(200, {
                    "taskId": task_id,
                    "snapshots": state.get_rank_history(task_id, keyword),
                })
                return

            if path == "/api/scrape/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                self._send_json(200, {"tasks": state.list_scrape_tasks()})
                return

            if path == "/api/scrape/products":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("taskId") or [""])[0]
                if not task_id:
                    self._send_json(400, {"error": "taskId required"})
                    return
                self._send_json(200, {
                    "taskId": task_id,
                    "products": state.get_scrape_products(task_id),
                })
                return

            if path == "/api/exports/list":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                self._send_json(200, {"jobs": state.list_export_jobs()})
                return

            if path == "/api/review/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                self._send_json(200, {"tasks": state.list_review_tasks()})
                return

            if path == "/api/review/results":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("taskId") or [""])[0]
                if not task_id:
                    self._send_json(400, {"error": "taskId required"})
                    return
                self._send_json(200, {
                    "taskId": task_id,
                    "results": state.get_review_results(task_id),
                })
                return

            if path == "/api/ai/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                skill_id = (q.get("skillId") or [None])[0]
                self._send_json(200, {"tasks": state.list_ai_tasks(skill_id)})
                return

            if path == "/api/ai/task":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                tid = (q.get("id") or [""])[0]
                if not tid:
                    self._send_json(400, {"error": "id required"})
                    return
                task = state.get_ai_task(tid)
                if not task:
                    self._send_json(404, {"error": "task not found"})
                    return
                self._send_json(200, task)
                return

            if path == "/api/ai/file":
                # 不做 Bearer 头校验：iframe/<img>/直接下载链接等浏览器原生导航不会带自定义头，
                # 和现有 /api/pdf/download 一致，靠不可猜测的 taskId + 文件名白名单做访问控制。
                q = parse_qs(urlparse(self.path).query)
                tid = (q.get("taskId") or [""])[0]
                name = (q.get("name") or [""])[0]
                mode = (q.get("mode") or ["download"])[0]
                task = state.get_ai_task(tid) if tid else None
                if not task:
                    self._send_json(404, {"error": "task not found"})
                    return
                allowed_names = {f["name"] for f in task.get("files", [])}
                if name not in allowed_names:
                    self._send_json(404, {"error": "file not found"})
                    return
                skill = AI_SKILLS.get(task["skill_id"])
                if not skill:
                    self._send_json(404, {"error": "skill not found"})
                    return
                fpath = os.path.join(skill["dir"], "report", tid, name)
                if not os.path.isfile(fpath):
                    self._send_json(404, {"error": "文件不存在"})
                    return
                self._send_file(fpath, download_name=name, inline=(mode == "inline"))
                return

            if path == "/api/pdf/download":
                q = parse_qs(urlparse(self.path).query)
                dl_id = (q.get("id") or [""])[0]
                if not dl_id:
                    self._send_json(400, {"error": "id required"})
                    return
                fpath = pdf_splitter.get_download_path(dl_id)
                if not fpath:
                    self._send_json(404, {"error": "文件不存在或已过期，请重新拆分"})
                    return
                self._send_file(fpath)
                return

            return super().do_GET()

        # ---- POST ----

        def do_POST(self):
            path = urlparse(self.path).path

            if path == "/api/login":
                payload = self._read_json()
                if payload is None:
                    return
                username = str(payload.get("username", "")).strip()
                password = str(payload.get("password", ""))
                token, user_info = auth.login(username, password)
                if token:
                    print(f"  [auth] {username} 登录成功")
                    self._send_json(200, {"token": token, "user": user_info})
                else:
                    self._send_json(401, {"error": "用户名或密码错误"})
                return

            if path == "/api/logout":
                token = _extract_token(self)
                if token:
                    auth.logout(token)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/rank/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                if not (payload.get("asin") or "").strip():
                    self._send_json(400, {"error": "asin required"})
                    return
                self._send_json(200, {"task": state.upsert_rank_task(payload)})
                return

            if path == "/api/rank/run":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                task_id = payload.get("taskId")
                task = state.get_rank_task(task_id) if task_id else payload
                if not task or not (task.get("asin") or "").strip():
                    self._send_json(400, {"error": "task not found / asin required"})
                    return
                if not task.get("id"):
                    task = state.upsert_rank_task(task)
                print(f"  [rank] 手动执行 {task['asin']}/{task['marketplace']}")
                results = run_rank_task(state, task)
                self._send_json(200, {"taskId": task["id"], "results": results})
                return

            if path == "/api/scrape/run":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                asins = payload.get("asins")
                if not isinstance(asins, list) or not asins:
                    self._send_json(400, {"error": "asins must be a non-empty array"})
                    return
                marketplace = (payload.get("marketplace") or "US").upper()
                if marketplace not in product_fetcher.MARKETPLACES:
                    self._send_json(400, {"error": f"unsupported marketplace: {marketplace}"})
                    return
                with_reviews = bool(payload.get("withReviews", False))
                task_id_in = payload.get("taskId")
                if not isinstance(task_id_in, str) or not task_id_in:
                    task_id_in = None
                total = payload.get("total")
                if not isinstance(total, int):
                    total = None
                print(f"  [scrape] 采集 {len(asins)} 个 ASIN @ {marketplace}"
                      f"{' (含评论)' if with_reviews else ''}")
                task_id, results = run_scrape_task(state, asins, marketplace, with_reviews, task_id=task_id_in, total=total)
                self._send_json(200, {"taskId": task_id, "results": results})
                return

            if path == "/api/review/run":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                asins = payload.get("asins")
                if not isinstance(asins, list) or not asins:
                    self._send_json(400, {"error": "asins must be a non-empty array"})
                    return
                marketplace = (payload.get("marketplace") or "US").upper()
                if marketplace not in product_fetcher.MARKETPLACES:
                    self._send_json(400, {"error": f"unsupported marketplace: {marketplace}"})
                    return
                sort_by = payload.get("sortBy") or "recent"
                if sort_by not in ("recent", "helpful"):
                    sort_by = "recent"
                filter_star = payload.get("filterStar")
                if filter_star is not None:
                    try:
                        filter_star = int(filter_star)
                    except (TypeError, ValueError):
                        filter_star = None
                    if filter_star not in (1, 2, 3, 4, 5):
                        filter_star = None
                verified_only = bool(payload.get("verifiedOnly", False))
                try:
                    max_pages = int(payload.get("maxPages", 3))
                except (TypeError, ValueError):
                    max_pages = 3
                max_pages = max(1, min(max_pages, 10))
                task_id_in = payload.get("taskId")
                if not isinstance(task_id_in, str) or not task_id_in:
                    task_id_in = None
                print(f"  [review] 采集 {len(asins)} 个 ASIN 评论 @ {marketplace}"
                      f" (sortBy={sort_by}, maxPages={max_pages})")
                task_id, results = run_review_task(
                    state, asins, marketplace, sort_by, filter_star, verified_only, max_pages,
                    task_id=task_id_in,
                )
                self._send_json(200, {"taskId": task_id, "results": results})
                return

            if path == "/api/ai/run":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                skill_id = (payload.get("skillId") or "").strip()
                if skill_id not in AI_SKILLS:
                    self._send_json(400, {"error": f"unsupported skillId: {skill_id}"})
                    return
                asin = (payload.get("asin") or "").strip().upper()
                if not re.match(r"^[A-Z0-9]{10}$", asin):
                    self._send_json(400, {"error": "asin 格式不正确（应为10位字母数字）"})
                    return
                params = payload.get("params") or {}
                tid = state.create_ai_task(skill_id, asin, params)
                print(f"  [ai] 创建分析任务 {tid}：{skill_id} / {asin}")
                self._send_json(200, {"taskId": tid})
                return

            if path == "/api/pdf/upload":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                if not pdf_splitter.check_pypdf():
                    self._send_json(503, {"error": "后端未安装 pypdf，请运行: pip3 install pypdf"})
                    return
                length = int(self.headers.get("Content-Length", "0"))
                if length == 0:
                    self._send_json(400, {"error": "empty body"})
                    return
                filename = unquote(self.headers.get("X-Filename", "upload.pdf"))
                data = self.rfile.read(length)
                try:
                    info = pdf_splitter.save_upload(filename, data)
                    self._send_json(200, info)
                except Exception as e:
                    self._send_json(500, {"error": str(e)})
                return

            if path == "/api/pdf/split":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                jobs = payload.get("jobs")
                if not isinstance(jobs, list) or not jobs:
                    self._send_json(400, {"error": "jobs must be a non-empty array"})
                    return
                results = [pdf_splitter.run_split_job(job) for job in jobs]
                self._send_json(200, {"results": results})
                return

            if path == "/api/pdf/zip":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                ids = payload.get("ids")
                if not isinstance(ids, list) or not ids:
                    self._send_json(400, {"error": "ids must be a non-empty array"})
                    return
                zip_name = str(payload.get("name", "split_results.zip"))
                try:
                    data = pdf_splitter.create_zip(ids)
                except Exception as e:
                    self._send_json(500, {"error": str(e)})
                    return
                from urllib.parse import quote
                encoded_name = quote(zip_name, safe="")
                self.send_response(200)
                self.send_header("Content-Type", "application/zip")
                self.send_header("Content-Length", str(len(data)))
                self.send_header(
                    "Content-Disposition",
                    f'attachment; filename="{zip_name.encode("ascii","replace").decode("ascii")}"; filename*=UTF-8\'\'{encoded_name}',
                )
                self.end_headers()
                self.wfile.write(data)
                return

            if path == "/api/exports/create":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                job_type = (payload.get("type") or "").strip()
                if job_type not in ("scrape_xlsx",):
                    self._send_json(400, {"error": f"unsupported type: {job_type}"})
                    return
                params  = payload.get("params") or {}
                label   = (payload.get("label") or "导出任务").strip()
                fname   = (payload.get("fileName") or f"export_{_now_iso()[:10]}.xlsx").strip()
                # 预估 progress_total（按产品数）
                task_id = params.get("taskId", "")
                asins   = params.get("asins")
                products = state.get_scrape_products(task_id) if task_id else []
                if asins:
                    products = [p for p in products if p.get("asin") in set(asins)]
                total = len(products)
                jid = state.create_export_job(job_type, label, {**params, "asins": asins}, total)
                state.update_export_job(jid, file_name=fname)
                self._send_json(200, {"jobId": jid})
                return

            if path == "/api/scrape/export-xlsx":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                task_id = (payload.get("taskId") or "").strip()
                if not task_id:
                    self._send_json(400, {"error": "taskId required"})
                    return
                products = state.get_scrape_products(task_id)
                if not products:
                    self._send_json(404, {"error": "该任务暂无产品数据"})
                    return
                try:
                    xlsx_bytes = _build_products_xlsx(products)
                except Exception as e:
                    self._send_json(500, {"error": str(e)})
                    return
                from urllib.parse import quote
                fname = f"amazon_products_{task_id[:8]}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self.send_header("Content-Disposition",
                    f'attachment; filename="{fname}"; filename*=UTF-8\'\'{quote(fname, safe="")}')
                self.end_headers()
                self.wfile.write(xlsx_bytes)
                return

            self.send_error(404)

        # ---- DELETE ----

        def do_DELETE(self):
            path = urlparse(self.path).path
            if path == "/api/rank/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("id") or [""])[0]
                if not task_id:
                    self._send_json(400, {"error": "id required"})
                    return
                state.delete_rank_task(task_id)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/rank/keyword":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("taskId") or [""])[0]
                keyword = (q.get("keyword") or [""])[0]
                if not task_id or not keyword:
                    self._send_json(400, {"error": "taskId and keyword required"})
                    return
                task = state.get_rank_task(task_id)
                if task:
                    new_kws = [k for k in task["keywords"] if k != keyword]
                    new_notes = {k: v for k, v in task["keywordNotes"].items() if k != keyword}
                    state.upsert_rank_task({**task, "keywords": new_kws, "keywordNotes": new_notes})
                    with state.lock:
                        with state._conn() as conn:
                            conn.execute(
                                "DELETE FROM rank_snapshots WHERE task_id=? AND keyword=?",
                                [task_id, keyword]
                            )
                            conn.commit()
                self._send_json(200, {"ok": True})
                return

            if path == "/api/scrape/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("id") or [""])[0]
                if not task_id:
                    self._send_json(400, {"error": "id required"})
                    return
                state.delete_scrape_task(task_id)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/review/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                task_id = (q.get("id") or [""])[0]
                if not task_id:
                    self._send_json(400, {"error": "id required"})
                    return
                state.delete_review_task(task_id)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/ai/tasks":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                tid = (q.get("id") or [""])[0]
                if not tid:
                    self._send_json(400, {"error": "id required"})
                    return
                task = state.get_ai_task(tid)
                if task:
                    skill = AI_SKILLS.get(task["skill_id"])
                    if skill:
                        import shutil
                        out_dir = os.path.join(skill["dir"], "report", tid)
                        shutil.rmtree(out_dir, ignore_errors=True)
                state.delete_ai_task(tid)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/scrape/reset-session":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                marketplace = (payload.get("marketplace") or "").upper()
                product_fetcher.reset_session(marketplace if marketplace else None)
                self._send_json(200, {"ok": True})
                return

            if path == "/api/exports":
                if not auth.verify(_extract_token(self)):
                    self._send_json(401, {"error": "请先登录"})
                    return
                q = parse_qs(urlparse(self.path).query)
                jid = (q.get("id") or [""])[0]
                if not jid:
                    self._send_json(400, {"error": "id required"})
                    return
                jobs = state.list_export_jobs()
                for j in jobs:
                    if j["id"] == jid and j.get("download_id"):
                        fpath = pdf_splitter._download_registry.pop(j["download_id"], None)
                        if fpath:
                            try:
                                os.remove(fpath)
                            except OSError:
                                pass
                        break
                state.delete_export_job(jid)
                self._send_json(200, {"ok": True})
                return

            self.send_error(404)

        # ---- PUT ----

        def do_PUT(self):
            path = urlparse(self.path).path

            if path == "/api/products":
                user = auth.verify(_extract_token(self))
                if not user:
                    self._send_json(401, {"error": "请先登录"})
                    return
                payload = self._read_json()
                if payload is None:
                    return
                new_products = payload.get("products")
                base_version = payload.get("baseVersion")
                if not isinstance(new_products, list):
                    self._send_json(400, {"error": "products must be an array"})
                    return
                new_version, conflict = state.write(
                    new_products, base_version, user=user.get("username")
                )
                if conflict is not None:
                    self._send_json(409, conflict)
                else:
                    self._send_json(200, {"version": new_version})
                return

            self.send_error(405, "PUT not allowed here")

        # ---- 工具方法 ----

        def _read_json(self):
            length = int(self.headers.get("Content-Length", "0"))
            raw = self.rfile.read(length).decode("utf-8")
            try:
                return json.loads(raw)
            except json.JSONDecodeError as e:
                self._send_json(400, {"error": "invalid JSON: " + str(e)})
                return None

        def _send_json(self, code, obj):
            body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
            self.send_response(code)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_file(self, path, download_name=None, inline=False):
            import mimetypes
            from urllib.parse import quote
            try:
                size = os.path.getsize(path)
                ctype, _ = mimetypes.guess_type(path)
                ctype = ctype or "application/octet-stream"
                fname = download_name or os.path.basename(path)
                # RFC 5987：非 ASCII 文件名用 UTF-8 percent-encoding
                ascii_fname = fname.encode("ascii", "replace").decode("ascii")
                encoded_fname = quote(fname, safe="")
                disposition = "inline" if inline else "attachment"
                self.send_response(200)
                self.send_header("Content-Type", ctype)
                self.send_header("Content-Length", str(size))
                self.send_header(
                    "Content-Disposition",
                    f'{disposition}; filename="{ascii_fname}"; filename*=UTF-8\'\'{encoded_fname}',
                )
                self.end_headers()
                with open(path, "rb") as f:
                    while True:
                        chunk = f.read(65536)
                        if not chunk:
                            break
                        self.wfile.write(chunk)
            except OSError:
                self.send_error(500)

    return Handler


class ThreadingServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads      = True
    allow_reuse_address = True


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description="FBA Tracker 局域网协作服务器 v2")
    p.add_argument("--port",    type=int, default=8002,           help="监听端口 (默认 8002)")
    p.add_argument("--host",    default="0.0.0.0",                help="监听地址 (默认 0.0.0.0)")
    p.add_argument("--db",      default="data/fba-data.db",       help="SQLite 数据库 (默认 ./data/fba-data.db)")
    p.add_argument("--users",   default="data/fba-users.json",    help="用户配置文件 (默认 ./data/fba-users.json)")
    p.add_argument("--migrate", default="data/fba-data.json",     help="首次启动时从旧 JSON 文件迁移 (默认 ./data/fba-data.json)")
    args = p.parse_args()

    db_path    = os.path.abspath(args.db)
    users_path = os.path.abspath(args.users)

    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    os.makedirs(os.path.dirname(users_path), exist_ok=True)

    db_state = DbState(db_path)
    db_state.import_from_json(os.path.abspath(args.migrate))
    pdf_splitter.cleanup_old_tmp()

    auth_mgr = AuthManager(users_path)

    export_worker = ExportWorker(db_state)
    export_worker.start()

    ai_worker = AiAnalysisWorker(db_state)
    ai_worker.start()

    handler  = make_handler(db_state, auth_mgr)
    server   = ThreadingServer((args.host, args.port), handler)

    ip        = get_lan_ip()
    local_hn  = socket.gethostname()  # e.g. Mac-miniY.local
    if not local_hn.endswith('.local'):
        local_hn = local_hn + '.local'
    bar = "─" * 58
    print(bar)
    print("  FBA Tracker 协作服务器 v2 已启动")
    print(bar)
    print(f"  本机访问：  http://localhost:{args.port}")
    print(f"  固定地址：  http://{local_hn}:{args.port}  ← 换 WiFi 也不变")
    print(f"  当前 IP：   http://{ip}:{args.port}")
    print(f"  数据库：    {db_path}")
    print(f"  用户配置：  {users_path}")
    print(f"  当前版本：  v{db_state.version}  ({db_state.product_count} 个产品)")
    print(bar)
    print("  固定地址（.local）在公司任意 WiFi 下均可访问")
    print("  Ctrl+C 停止服务")
    print(bar)
    start_scheduler(db_state)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止")


if __name__ == "__main__":
    main()
